import cx_Oracle
import json 
import google.generativeai as genai
from sklearn.feature_extraction.text import CountVectorizer
from pathlib import Path
from dotenv import load_dotenv
import os
import re 
import csv
import pandas as pd
# import joblib

import os
import openpyxl
import random
import datetime
import nepali_datetime

load_dotenv()
GOOGLE_API_KEY=os.getenv('GOOGLE_API_KEY')
print(GOOGLE_API_KEY)



# class a_Model:
#     def __init__(self, db_user, db_pass, host, port, service):
#         dsn_tns = cx_Oracle.makedsn(host, port, service_name=service)
#         self.conn = cx_Oracle.connect(
#             user=db_user, password=db_pass, dsn=dsn_tns)
#         print('user name is ',db_user)
#         print('password is', db_pass)
#         print('Service name', service)

#         self.cur = self.conn.cursor()
#         self.db_user = db_user
#         # self.model = joblib.load('model.pkl')  #new


#         nepali_month_names = [
#                     "baishakh", "jestha", "ashadh", "shrawan", "bhadra",
#                     "ashoj", "kartik", "mangsir", "poush", "magh",
#                     "falgun", "chaitra"
#                 ]

#         today_date = datetime.date.today()
#         today_nepali = nepali_datetime.date.today()

#         self.today_date = today_date
#         self.today_nepali = today_nepali  

#         yesterday_date = today_date - datetime.timedelta(days=1)

#         if today_nepali.day == 1:
#             if today_nepali.month == 1:  
#                 self.yesterday_nepali = nepali_datetime.date(today_nepali.year - 1, 12, 30)  
#             else:
#                 last_month = today_nepali.month - 1
#                 days_in_last_month = nepali_datetime.date(today_nepali.year, last_month, 1).days_in_month
#                 self.yesterday_nepali = nepali_datetime.date(today_nepali.year, last_month, days_in_last_month)
#         else:
#             self.yesterday_nepali = nepali_datetime.date(today_nepali.year, today_nepali.month, today_nepali.day - 1)

#         seven_days_back_nepali = today_nepali - datetime.timedelta(days=7)

#         if seven_days_back_nepali.day == 1:
#             if seven_days_back_nepali.month == 1:
#                 seven_days_back_nepali = nepali_datetime.date(seven_days_back_nepali.year - 1, 12, 30)
#             else:
#                 last_month = seven_days_back_nepali.month - 1
#                 days_in_last_month = nepali_datetime.date(seven_days_back_nepali.year, last_month, 1).days_in_month
#                 seven_days_back_nepali = nepali_datetime.date(seven_days_back_nepali.year, last_month, days_in_last_month)

#         self.seven_days_back_nepali = seven_days_back_nepali

#         self.this_nepali_month = nepali_month_names[today_nepali.month - 1]  

#         if today_nepali.month == 1:  
#             self.last_nepali_month = "Chaitra"
#         else:
#             # -2 is done because to name previsous month form the above fiven list: nepali_moth_names
#             self.last_nepali_month = nepali_month_names[today_nepali.month - 2]
       
#         print("*********************************************************")

#         fiscal_year = self.db_user
#         self.fiscal_year = fiscal_year[-4:]
#         print(self.fiscal_year)
        
#         if self.fiscal_year == '8081':
#             self.filepath = 'dataset_8081.xlsx'

#         if self.fiscal_year == '8182':
#             self.filepath = 'dataset_8182.xlsx'
#             print("hellolo")

#         print(self.filepath)

#     def detect_query_type(self, query: str): #new
#         """
#         Uses the trained Naive Bayes model to predict whether the query is related to stock or sales.
#         """
#         # Vectorize the query using the same vectorizer used during model training
#         vecterizer = CountVectorizer()
#         query_vector = vecterizer.get_transform([query])

#         # Predict using the Naive Bayes model
#         prediction = self.model.predict(query_vector)

#         # # Return the predicted class: 'stock' or 'sales'
#         # if prediction == 0:
#         #     return 'stock'
#         # elif prediction == 1:
#         #     return 'sales'
#         # else:
#         return prediction   #new


#     def fetch_query(self, txt,query_detect, company_code):
#         txt = txt.lower()
#         GOOGLE_API_KEY = 'AIzaSyApeWHTR0BcNsr4tmh3IqQ2KOsf2ICY0Js'
#         genai.configure(api_key=GOOGLE_API_KEY)

#         if query_detect =="sales":
#             generation_config = {
#             "temperature": 0.8,
#             "top_p": 0.95,
#             "top_k": 64,
#             "max_output_tokens": 8192,
#             "response_mime_type": "text/plain",
#             }


#             #Prompt file is opened and replacing dates. This prompt is stored in the variable system_instruction.
#             filename = 'sales_prompt.txt'
#             with open(filename, 'r') as file:
#                 system_instruction = file.read()
#                 print(self.last_nepali_month)
#                 print(self.today_nepali)
#                 print(self.this_nepali_month)
#                 system_instruction = system_instruction.replace('yesterday_date',str(self.yesterday_nepali))
#                 system_instruction = system_instruction.replace('today_date',str(self.today_nepali))
#                 system_instruction = system_instruction.replace('last_nepali_month',str(self.last_nepali_month))
#                 system_instruction = system_instruction.replace('this_month',str(self.this_nepali_month))
#                 system_instruction = system_instruction.replace('seven_days_back_nepali_date',str(self.seven_days_back_nepali))
#                 if self.fiscal_year == '8182':
#                     system_instruction = system_instruction.replace("2081","2082")

#             #System Instruction is passed to the LLM as a prompt
#             model = genai.GenerativeModel(
#             model_name="gemini-1.5-flash",
#             generation_config=generation_config,
#             system_instruction=system_instruction,)
            
#             file_path = self.filepath
#             df = pd.read_excel(file_path)

#             df['input_lower'] = df['input'].str.lower()

#             user_input_lower = txt.lower()
#             question = txt
#             #At first checks the user prompt to the excel file
#             if user_input_lower in df['input_lower'].values:
#                 response = df.loc[df['input_lower'] == user_input_lower, 'output'].values[0]
#             else:
#                 #If not present then generate it through LLM
#                 print(txt)
#                 response = model.generate_content(txt)
#                 response = response.text
#             output = self.read_sql_query(response, question)
#             return output

#         if query_detect =="stock":
#             generation_config = {
#             "temperature": 1,
#             "top_p": 0.95,
#             "top_k": 64,
#             "max_output_tokens": 8192,
#             "response_mime_type": "text/plain",
#             }

#             filename = 'stock_prompt.txt'
#             with open(filename, 'r') as file:
#                 system_instruction = file.read()

#             model = genai.GenerativeModel(
#             model_name="gemini-1.5-flash",
#             generation_config=generation_config,
#             system_instruction= system_instruction,)
            
#             file_path = self.filepath
#             df = pd.read_excel(file_path)

#             df['input_lower'] = df['input'].str.lower()

#             user_input_lower = txt.lower()
#             question = txt
#             if user_input_lower in df['input_lower'].values:
#                 response = df.loc[df['input_lower'] == user_input_lower, 'output'].values[0]
#             else:
#                 response = model.generate_content(txt)
#                 response = response.text
#             output = self.read_sql_query(response, question)
#             return output

#     def read_sql_query(self, qry, question):
#         #Process a SQL query, execute it against a database and returns a result in a structured format"
        
#         #Cleaning SQL queries by replacing terminator, replacing if sql is present, backticks nad leading and trailing spaces
#         qry = qry.replace(";", "")
#         qry = qry.replace("sql", "")
#         qry = qry.replace("```", "")
#         qry = qry.strip() 

#         #Captures current date and time in ISO format. It is useful for logging
#         current_datetime = nepali_datetime.datetime.now().isoformat()
#         #creating dictionary for logging
#         log_entry = {
#             "datetime": current_datetime,
#             "input": question,
#             "output": qry
#         }
#         filename = self.filepath

#         #relative dates
#         pattern = r'\b(last|yesterday|today|today\'s|past|since|lastmonth|lastweek|thismonth|thisweek|nextmonth|nextweek|previous|previous month|previous week|current|current month|current week|after|before|ago|tomorrow|daybeforeyesterday|dayaftertomorrow|recent|recently|in the past|in the last|in the previous|this month|this week|this year|this month\'s|this week\'s|this year\'s|today\'s|yesterday\'s|tomorrow\'s)\b'

#         #Check is the query Contains a Relative Date
#         if re.search(pattern, log_entry['input'].lower()):
#             print("------------------Input contains a relative date keyword; entry will not be logged.")
#         else:
            
#             #If the query doesn't contain a relative date, it logs the details to an Excel file. If the file exists, it loads the workbook; otherwise, it creates a new one with headers for datetime, input, and output.
#             if os.path.exists(filename):
#                 workbook = openpyxl.load_workbook(filename)
#                 sheet = workbook.active
#             else:
#                 workbook = openpyxl.Workbook()
#                 sheet = workbook.active
#                 sheet.append(["datetime", "input", "output"])

#             next_row = sheet.max_row + 1

#             sheet[f'A{next_row}'] = log_entry['datetime']
#             sheet[f'B{next_row}'] = log_entry['input']
#             sheet[f'C{next_row}'] = log_entry['output']

#             workbook.save(filename)


#         print(qry)

#         #It attempts to execute an SQL query using self.cur.execute(), instance of databse cursor
#         try:
#             self.cur.execute(qry)
#             #fetches result
#             all_codes = self.cur.fetchall()
#         #if fetch doesnot happen then error occured.
#         except Exception as e:
#             print(e)
#             return {"data": {}, "query": qry, "question": question}
#         finally:
#             pass
#         # retrives column name from the query result, cur.description contains the metadata about the columns returned by the query
#         all_cols = [desc[0] for desc in self.cur.description]
#         print("----------------------------------------------")

#         # Now organizing data in dictionary format
#         #For each column (iterated through all_cols), it creates a list (asd) containing the values for that column from all rows in the all_codes result set.
#         data = {}
#         for z in range(len(all_cols)):
#             asd = []
#             for i in all_codes:
#                 asd.append(str(i[z]))
#             data[all_cols[z]] = asd
#         #for creating graph. It pairs first column with each other column which might be sueful for creating graph axes or time series
#         graph_keys = []
#         for i in range(1, len(all_cols)):
#             graph_keys.append([all_cols[0], all_cols[i]])

#         # for creating a random response
#         desc = self.get_random_response()
#         return {"data":data, "query":qry, "question": question, "graph_keys":graph_keys, "desc":desc}

#     def get_random_response(self):
#         responses = [
#             "Here's your freshly compiled list.",
#             "Your requested summary is ready below.",
#             "The latest data you asked for is now available.",
#             "Here's the detailed response you requested.",
#             "Your list has been generated and is displayed below.",
#             "We've put together your summary; check it out below.",
#             "Below is the response we've prepared for you.",
#             "Your results are in! See your data below.",
#             "The analyzed data is ready for you; find it below.",
#             "Here's the summary you asked for, ready and waiting below.",
#             "Your list is ready! Check out the details below.",
#             "We've gathered the information you needed; view your summary below.",
#             "Here's your customized list, freshly prepared.",
#             "The response you requested is now available below.",
#             "Your latest data insights are ready; see them below.",
#             "Below is the detailed summary you've been waiting for.",
#             "We've generated the data you needed; review it below.",
#             "Your detailed list is prepared and ready for you.",
#             "Find your compiled response below, ready for review.",
#             "The results you requested are now available below.",
#             "Check out your generated summary, displayed below.",
#             "Here's the full data you asked for, ready to view.",
#             "The list you need is compiled and shown below.",
#             "Your response is complete and available for review below.",
#             "We've finalized your summary; find it below.",
#             "Below is the comprehensive list you requested.",
#             "Your latest response is ready; see the details below.",
#             "Here's your complete data, freshly generated.",
#             "The analysis you requested has been prepared; view it below.",
#             "Your list has been successfully compiled; check it out below.",
#             "Find your requested details in the response displayed below.",
#             "The summary is ready and available for you below.",
#             "Here's the detailed data, generated just for you.",
#             "We've assembled the response you needed; find it below.",
#             "Your requested data is ready; view the list below.",
#             "Below is the final response you asked for, ready for your review.",
#             "Your customized list is now available below.",
#             "Here's the analysis you requested, displayed below.",
#             "Your summary has been prepared; see it below.",
#             "The results you needed are ready and waiting below."
#         ]
#         desc = random.choice(responses)
#         return desc
    



    